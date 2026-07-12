"""
Builds dev-cost-estimate.xlsx from the module estimates in 08-roadmap.md.
Formula-driven: change day-rate or team scenario multipliers and totals recompute.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
INPUT_FILL = PatternFill("solid", fgColor="FFF9DB")
INPUT_FONT = Font(color="7A5C00")
CALC_FONT = Font(color="1D4ED8")
TITLE_FONT = Font(bold=True, size=14)
NOTE_FONT = Font(italic=True, size=9, color="6B7280")
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(ws, row, c1, c2):
    for c in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

ws = wb.active
ws.title = "Assumptions"
ws["A1"] = "Dev Cost Estimate — Assumptions (жёлтое — редактируемо)"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:D1")

rows = [
    ("Effective AI-first дней в месяце на 1 разработчика", 20, "дней"),
    ("Дневная ставка AI-first разработчика (S1/S2, включая AI-инструменты)", 220, "USD/день, ориентир на РФ/СНГ remote-рынок 2026"),
    ("Дневная ставка для S3 (команда, средневзвешенная по ролям)", 200, "USD/день"),
    ("Множитель скорости S1 (соло)", 1.6, "x — параллельные AI-агенты, без издержек человеческой синхронизации"),
    ("Множитель скорости S2 (2 разработчика)", 1.0, "x — издержки координации между людьми перевешивают выигрыш от 2-й пары рук"),
    ("Множитель скорости S3 (команда 4-5 чел.)", 2.8, "x"),
    ("Кол-во кодящих FTE в S3", 2.8, "FTE-эквивалент, см. 08-roadmap.md"),
]
r = 3
A = {}
for label, val, unit in rows:
    ws.cell(row=r, column=1, value=label).border = BORDER
    c = ws.cell(row=r, column=2, value=val)
    c.fill = INPUT_FILL
    c.font = INPUT_FONT
    c.border = BORDER
    ws.cell(row=r, column=3, value=unit).font = NOTE_FONT
    A[label] = f"Assumptions!$B${r}"
    r += 1
autosize(ws, [55, 12, 45])

ws2 = wb.create_sheet("Modules")
ws2["A1"] = "Смета по модулям (AI-дни, соло-эквивалент) — источник: 08-roadmap.md §8.2"
ws2["A1"].font = TITLE_FONT
ws2.merge_cells("A1:F1")

headers = ["Модуль", "Мин. (AI-дни)", "Реалистично (AI-дни)", "Пессимистично (AI-дни)",
           "Мин. (USD, S1)", "Реалистично (USD, S1)", "Пессимистично (USD, S1)"]
hr = 3
for i, h in enumerate(headers, start=1):
    ws2.cell(row=hr, column=i, value=h)
style_header(ws2, hr, 1, len(headers))

modules = [
    ("Инфраструктура проекта (monorepo, CI/CD, окружения)", 2, 4, 7),
    ("Auth (Telegram initData + опц. web OTP)", 2, 4, 8),
    ("Wallet & Billing ledger (hold/commit/refund, idempotency)", 3, 6, 12),
    ("Платёжная интеграция (Midtrans/Xendit, QRIS, e-wallets)", 3, 6, 12),
    ("Model Registry / Catalog", 1, 3, 5),
    ("AI Gateway (унифицированный клиент, ретраи, фолбэк, cost logging)", 4, 8, 15),
    ("Generation API + Queue + Workers (image)", 3, 6, 10),
    ("Generation API + Queue + Workers (video)", 5, 10, 18),
    ("Generation (audio/voice)", 2, 4, 8),
    ("Чат/LLM интеграция (мультимодельный чат)", 2, 5, 9),
    ("Telegram Mini App (UI, генерация, история, баланс)", 5, 10, 18),
    ("Telegram Bot (команды, уведомления, webhook)", 2, 4, 7),
    ("История генераций, медиатека, шаринг", 2, 4, 7),
    ("Admin Panel", 2, 5, 9),
    ("Observability (логи, метрики, алерты)", 2, 4, 6),
    ("Аналитика продукта", 1, 3, 5),
    ("Локализация (Bahasa Indonesia + English)", 1, 3, 5),
    ("Compliance/юр. (ToS, Privacy, UU PDP, возрастные ограничения)", 2, 4, 8),
    ("Модерация контента (NSFW-фильтр)", 2, 4, 8),
    ("QA / нагрузочное тестирование / security review", 2, 5, 9),
]

r0 = hr + 1
for i, (name, mn, re_, pe) in enumerate(modules):
    rr = r0 + i
    ws2.cell(row=rr, column=1, value=name).border = BORDER
    ws2.cell(row=rr, column=2, value=mn).border = BORDER
    ws2.cell(row=rr, column=3, value=re_).border = BORDER
    ws2.cell(row=rr, column=4, value=pe).border = BORDER
    for col, src_col in zip([5, 6, 7], [2, 3, 4]):
        col_l = get_column_letter(col)
        src_l = get_column_letter(src_col)
        cell = ws2.cell(row=rr, column=col)
        cell.value = f"={src_l}{rr}*{A['Дневная ставка AI-first разработчика (S1/S2, включая AI-инструменты)']}"
        cell.font = CALC_FONT
        cell.number_format = "#,##0"
        cell.border = BORDER

total_row = r0 + len(modules)
ws2.cell(row=total_row, column=1, value="ИТОГО").font = Font(bold=True)
for col in range(2, 8):
    col_l = get_column_letter(col)
    cell = ws2.cell(row=total_row, column=col)
    cell.value = f"=SUM({col_l}{r0}:{col_l}{total_row-1})"
    cell.font = Font(bold=True, color="1D4ED8")
    if col >= 5:
        cell.number_format = "#,##0"
    ws2.cell(row=total_row, column=col).border = BORDER
ws2.cell(row=total_row, column=1).border = BORDER

autosize(ws2, [52, 12, 12, 14, 12, 14, 16])
ws2.freeze_panes = "A4"

ws3 = wb.create_sheet("Team Scenarios")
ws3["A1"] = "Пересчёт на команды и календарные сроки"
ws3["A1"].font = TITLE_FONT
ws3.merge_cells("A1:E1")

hdr = ["Сценарий", "Множитель скорости", "Мин. (мес.)", "Реалистично (мес.)", "Пессимистично (мес.)",
       "Мин. (USD)", "Реалистично (USD)", "Пессимистично (USD)"]
hr3 = 3
for i, h in enumerate(hdr, start=1):
    ws3.cell(row=hr3, column=i, value=h)
style_header(ws3, hr3, 1, len(hdr))

scenarios = [
    ("S1 — 1 разработчик", A['Множитель скорости S1 (соло)'],
     f"Modules!$B${total_row}", f"Modules!$C${total_row}", f"Modules!$D${total_row}",
     A['Дневная ставка AI-first разработчика (S1/S2, включая AI-инструменты)']),
    ("S2 — 2 разработчика", A['Множитель скорости S2 (2 разработчика)'],
     f"Modules!$B${total_row}", f"Modules!$C${total_row}", f"Modules!$D${total_row}",
     A['Дневная ставка AI-first разработчика (S1/S2, включая AI-инструменты)']),
    ("S3 — команда 4-5 чел.", A['Множитель скорости S3 (команда 4-5 чел.)'],
     f"Modules!$B${total_row}", f"Modules!$C${total_row}", f"Modules!$D${total_row}",
     A['Дневная ставка для S3 (команда, средневзвешенная по ролям)']),
]

r3 = hr3 + 1
for name, mult, days_min, days_re, days_pe, rate in scenarios:
    ws3.cell(row=r3, column=1, value=name).border = BORDER
    ws3.cell(row=r3, column=2, value=f"={mult}").border = BORDER
    ws3.cell(row=r3, column=2).font = CALC_FONT
    days_per_month = A['Effective AI-first дней в месяце на 1 разработчика']
    for col, days in zip([3, 4, 5], [days_min, days_re, days_pe]):
        cell = ws3.cell(row=r3, column=col)
        cell.value = f"=({days})/({days_per_month}*{mult})"
        cell.font = CALC_FONT
        cell.number_format = "0.0"
        cell.border = BORDER
    for col, days in zip([6, 7, 8], [days_min, days_re, days_pe]):
        cell = ws3.cell(row=r3, column=col)
        cell.value = f"=({days})*({rate})"
        cell.font = CALC_FONT
        cell.number_format = "#,##0"
        cell.border = BORDER
    r3 += 1

ws3.cell(row=r3 + 1, column=1,
         value="Методология оценки дней/модуль — см. docs/due-diligence/syntx/08-roadmap.md §8.1-8.2 (откалибровано по истории коммитов этого репозитория).")
ws3.cell(row=r3 + 1, column=1).font = NOTE_FONT
ws3.merge_cells(start_row=r3 + 1, start_column=1, end_row=r3 + 1, end_column=6)

autosize(ws3, [26, 14, 12, 14, 16, 12, 16, 18])

wb.save("dev-cost-estimate.xlsx")
print("Saved dev-cost-estimate.xlsx")
