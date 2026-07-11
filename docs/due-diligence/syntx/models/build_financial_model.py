"""
Builds financial-model.xlsx — a living, formula-driven financial model.
All business assumptions live in the "Assumptions" sheet; every other sheet
references those cells so changing an assumption cascades everywhere (per
brief: "Параметры должны изменяться").

Run: python3 build_financial_model.py
Requires: openpyxl
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference

wb = openpyxl.Workbook()

# ---------- styling helpers ----------
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SUBHEADER_FILL = PatternFill("solid", fgColor="E5E7EB")
SUBHEADER_FONT = Font(bold=True, size=10)
INPUT_FILL = PatternFill("solid", fgColor="FFF9DB")
INPUT_FONT = Font(color="7A5C00")
CALC_FONT = Font(color="1D4ED8")
TITLE_FONT = Font(bold=True, size=14)
NOTE_FONT = Font(italic=True, size=9, color="6B7280")
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(ws, row, col1, col2):
    for c in range(col1, col2 + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# =====================================================================
# SHEET 1: Assumptions (all inputs, yellow = editable)
# =====================================================================
ws = wb.active
ws.title = "Assumptions"
ws["A1"] = "Nusa AI — Financial Model: Assumptions (все жёлтые ячейки редактируемые)"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:D1")

row = 3
sections = [
    ("PRICING (USD/mo, привязано к Rupiah ориентировочно)", [
        ("Free tier — token allowance/mo", 150, "токены"),
        ("Tier Basic — price", 3.5, "USD/mo (~Rp 55,000)"),
        ("Tier Basic — token allowance/mo", 1000, "токены"),
        ("Tier Pro — price", 9.0, "USD/mo (~Rp 145,000)"),
        ("Tier Pro — token allowance/mo", 3500, "токены"),
        ("Tier Business — price", 24.0, "USD/mo (~Rp 385,000)"),
        ("Tier Business — token allowance/mo", 12000, "токены"),
        ("Top-up pack price (extra 1000 tokens)", 3.0, "USD"),
    ]),
    ("USER MIX / CONVERSION", [
        ("% free users who convert to any paid tier /mo", 0.04, "доля"),
        ("Paid mix — % on Basic", 0.55, "доля от платящих"),
        ("Paid mix — % on Pro", 0.35, "доля от платящих"),
        ("Paid mix — % on Business", 0.10, "доля от платящих"),
        ("Monthly churn — paid users", 0.09, "доля/мес"),
        ("Avg top-up purchases per paying user/mo", 0.3, "шт"),
    ]),
    ("COGS — стоимость генерации (USD), см. Этап 04/05", [
        ("Avg cost per image generation (blended)", 0.012, "USD, см. 04.2 (Runware/DeepInfra/Imagen blend)"),
        ("Avg cost per video generation (5s, blended)", 0.35, "USD, см. 04.3 (Wan/Kling/Veo blend)"),
        ("Avg cost per chat message (LLM, blended)", 0.0025, "USD, см. 04.1 (Haiku/Flash-Lite blend)"),
        ("Avg cost per audio/voice generation", 0.02, "USD, см. 04.4 (ElevenLabs/Groq blend)"),
        ("Tokens charged per image generation", 8, "токены — продуктовое решение, не факт SYNTX"),
        ("Tokens charged per video generation (5s)", 120, "токены"),
        ("Tokens charged per chat message", 1, "токены"),
        ("Tokens charged per audio generation", 15, "токены"),
        ("Generation mix — % of TOKENS spent on image", 0.55, "доля токен-трат"),
        ("Generation mix — % of TOKENS spent on video", 0.15, "доля токен-трат"),
        ("Generation mix — % of TOKENS spent on chat", 0.25, "доля токен-трат"),
        ("Generation mix — % of TOKENS spent on audio", 0.05, "доля токен-трат"),
        ("Allowance utilization rate (breakage — доля фактически потраченных токенов от пакета)", 0.55, "доля, [EST] типично 40-70% для credit-моделей"),
        ("Payment processing fee (Midtrans/Xendit avg)", 0.029, "доля от revenue"),
    ]),
    ("CAC / MARKETING", [
        ("Blended CAC per new registered user", 0.9, "USD (Telegram Ads, influencers, ASO)"),
        ("Blended CAC per new PAYING user", 14.0, "USD"),
        ("Organic/referral share of new users", 0.35, "доля, снижает blended CAC"),
    ]),
    ("OPEX — фиксированные ежемесячные расходы (USD)", [
        ("Team cost (S1: 1 AI-first разработчик + fractional legal/design)", 5500, "USD/мес — $220/день x 20 дней (см. 08) + ~$1,100 на подрядчиков"),
        ("Infra fixed (hosting, DB, CDN, monitoring, GPU reserved)", 900, "USD/мес при <10k users"),
        ("Infra fixed scaling step (+ per 10k users)", 350, "USD/мес добавка на каждые 10k MAU"),
        ("Content moderation / compliance tooling", 300, "USD/мес"),
        ("Office/admin/legal overhead", 600, "USD/мес"),
    ]),
    ("ФИНАНСИРОВАНИЕ / БАЗА", [
        ("Стартовый капитал (seed)", 150000, "USD — нижняя граница диапазона из ТЗ"),
        ("Курс USD/IDR (справочно)", 16000, "IDR за 1 USD, июль 2026"),
    ]),
]

input_cells = {}
for title, rows in sections:
    ws.cell(row=row, column=1, value=title)
    style_header(ws, row, 1, 4)
    row += 1
    for label, val, unit in rows:
        ws.cell(row=row, column=1, value=label).border = BORDER
        c = ws.cell(row=row, column=2, value=val)
        c.fill = INPUT_FILL
        c.font = INPUT_FONT
        c.border = BORDER
        ws.cell(row=row, column=3, value=unit).font = NOTE_FONT
        input_cells[label] = f"Assumptions!$B${row}"
        row += 1
    row += 1

autosize(ws, [55, 14, 40, 5])
ws.freeze_panes = "A2"

# expose a python dict of cell refs for formulas in later sheets
A = input_cells

# =====================================================================
# SHEET 2: Unit Economics
# =====================================================================
ws2 = wb.create_sheet("Unit Economics")
ws2["A1"] = "Unit Economics (расчёт по формулам из Assumptions)"
ws2["A1"].font = TITLE_FONT
ws2.merge_cells("A1:D1")

row_labels = [
    "Blended ARPU (paying user, /mo)",
    "Avg token allowance (weighted by tier mix, /mo)",
    "Tokens actually consumed /mo (allowance x breakage/utilization)",
    "Blended $/token cost (derived from per-generation-type cost & tokens/gen, weighted by token-spend mix)",
    "Blended COGS per paying user (/mo) = tokens consumed x $/token + payment fee",
    "Gross margin per paying user (USD/mo)",
    "Gross margin %",
    "Monthly churn (paid)",
    "Customer lifetime (months)",
    "LTV (gross-margin based)",
    "CAC per paying user (blended incl. organic)",
    "LTV / CAC",
    "Payback period (months)",
]
r = 3
for label in row_labels:
    ws2.cell(row=r, column=1, value=label).border = BORDER
    ws2.cell(row=r, column=2).border = BORDER
    r += 1

ws2["B3"] = (
    f"=({A['Tier Basic — price']}*{A['Paid mix — % on Basic']}+{A['Tier Pro — price']}*{A['Paid mix — % on Pro']}+{A['Tier Business — price']}*{A['Paid mix — % on Business']})"
    f"+({A['Top-up pack price (extra 1000 tokens)']}*{A['Avg top-up purchases per paying user/mo']})"
)
ws2["B4"] = (
    f"=({A['Tier Basic — token allowance/mo']}*{A['Paid mix — % on Basic']}+{A['Tier Pro — token allowance/mo']}*{A['Paid mix — % on Pro']}+{A['Tier Business — token allowance/mo']}*{A['Paid mix — % on Business']})"
)
ws2["B5"] = f"=B4*{A['Allowance utilization rate (breakage — доля фактически потраченных токенов от пакета)']}"
ws2["B6"] = (
    f"=({A['Avg cost per image generation (blended)']}/{A['Tokens charged per image generation']})*{A['Generation mix — % of TOKENS spent on image']}"
    f"+({A['Avg cost per video generation (5s, blended)']}/{A['Tokens charged per video generation (5s)']})*{A['Generation mix — % of TOKENS spent on video']}"
    f"+({A['Avg cost per chat message (LLM, blended)']}/{A['Tokens charged per chat message']})*{A['Generation mix — % of TOKENS spent on chat']}"
    f"+({A['Avg cost per audio/voice generation']}/{A['Tokens charged per audio generation']})*{A['Generation mix — % of TOKENS spent on audio']}"
)
ws2["B7"] = f"=B5*B6+B3*{A['Payment processing fee (Midtrans/Xendit avg)']}"
ws2["B8"] = "=B3-B7"
ws2["B9"] = "=B8/B3"
ws2["B10"] = f"={A['Monthly churn — paid users']}"
ws2["B11"] = "=1/B10"
ws2["B12"] = "=B8*B11"
ws2["B13"] = f"={A['Blended CAC per new PAYING user']}*(1-{A['Organic/referral share of new users']})"
ws2["B14"] = "=B12/B13"
ws2["B15"] = "=B13/B8"

for rr in range(3, 16):
    ws2.cell(row=rr, column=2).font = CALC_FONT
    ws2.cell(row=rr, column=2).number_format = "0.0000" if rr == 6 else "0.00"
ws2["B9"].number_format = "0.0%"
ws2["B10"].number_format = "0.0%"

ws2["A17"] = "Примечание: ARPU/COGS считаются по paying-пользователю. Free-пользователи учтены в Scenarios через conversion rate. Tokens/generation — продуктовое допущение (не факт SYNTX, см. Этап 06), выбрано так, чтобы блендед $/token был внутренне согласован со стоимостями из Этапа 04."
ws2["A17"].font = NOTE_FONT
ws2.merge_cells("A17:E17")
autosize(ws2, [48, 16])

UE = {
    "ARPU": "'Unit Economics'!$B$3",
    "COGS_per_user": "'Unit Economics'!$B$7",
    "GM_per_user": "'Unit Economics'!$B$8",
    "GM_pct": "'Unit Economics'!$B$9",
    "Churn": "'Unit Economics'!$B$10",
    "Lifetime": "'Unit Economics'!$B$11",
    "LTV": "'Unit Economics'!$B$12",
    "CAC": "'Unit Economics'!$B$13",
    "LTV_CAC": "'Unit Economics'!$B$14",
    "Payback": "'Unit Economics'!$B$15",
}

# =====================================================================
# SHEET 3: Scenarios (MAU ladder: 100 .. 100,000)
# =====================================================================
ws3 = wb.create_sheet("Scenarios")
ws3["A1"] = "Scenarios by MAU (Monthly Active Users) — все метрики считаются формулами"
ws3["A1"].font = TITLE_FONT
ws3.merge_cells("A1:I1")

mau_values = [100, 500, 1000, 5000, 10000, 50000, 100000]
headers = ["Метрика"] + [f"{m:,} MAU" for m in mau_values]
ws3.append([])
hr = 3
for i, h in enumerate(headers, start=1):
    ws3.cell(row=hr, column=i, value=h)
style_header(ws3, hr, 1, len(headers))

metric_rows = [
    "MAU",
    "Paying users (= MAU * conversion)",
    "MRR (USD)",
    "ARR (USD)",
    "COGS total (USD/mo)",
    "Gross profit (USD/mo)",
    "Gross margin %",
    "Opex fixed (USD/mo)",
    "EBITDA (USD/mo)",
    "EBITDA margin %",
    "Burn rate (USD/mo, if EBITDA<0)",
    "CAC spend needed for new payers this mo (USD, assume 15% of payers are net-new/mo)",
    "LTV (USD)",
    "LTV/CAC",
    "Payback (months)",
    "Break-even MAU (paying-equivalent, static at this Opex level)",
]

start_row = hr + 1
for i, m in enumerate(metric_rows):
    ws3.cell(row=start_row + i, column=1, value=m).border = BORDER

col_letters = [get_column_letter(2 + i) for i in range(len(mau_values))]

for i, (mau, col) in enumerate(zip(mau_values, col_letters)):
    r_mau = start_row
    r_pay = start_row + 1
    r_mrr = start_row + 2
    r_arr = start_row + 3
    r_cogs = start_row + 4
    r_gp = start_row + 5
    r_gm = start_row + 6
    r_opex = start_row + 7
    r_ebitda = start_row + 8
    r_ebitdam = start_row + 9
    r_burn = start_row + 10
    r_cac = start_row + 11
    r_ltv = start_row + 12
    r_ltvcac = start_row + 13
    r_payback = start_row + 14
    r_be = start_row + 15

    ws3[f"{col}{r_mau}"] = mau
    ws3[f"{col}{r_pay}"] = f"={col}{r_mau}*{A['% free users who convert to any paid tier /mo']}"
    ws3[f"{col}{r_mrr}"] = f"={col}{r_pay}*{UE['ARPU']}"
    ws3[f"{col}{r_arr}"] = f"={col}{r_mrr}*12"
    ws3[f"{col}{r_cogs}"] = f"={col}{r_pay}*{UE['COGS_per_user']}"
    ws3[f"{col}{r_gp}"] = f"={col}{r_mrr}-{col}{r_cogs}"
    ws3[f"{col}{r_gm}"] = f"={col}{r_gp}/{col}{r_mrr}"
    ws3[f"{col}{r_opex}"] = (
        f"={A['Team cost (S1: 1 AI-first разработчик + fractional legal/design)']}"
        f"+{A['Infra fixed (hosting, DB, CDN, monitoring, GPU reserved)']}"
        f"+{A['Infra fixed scaling step (+ per 10k users)']}*({col}{r_mau}/10000)"
        f"+{A['Content moderation / compliance tooling']}"
        f"+{A['Office/admin/legal overhead']}"
    )
    ws3[f"{col}{r_ebitda}"] = f"={col}{r_gp}-{col}{r_opex}"
    ws3[f"{col}{r_ebitdam}"] = f"={col}{r_ebitda}/{col}{r_mrr}"
    ws3[f"{col}{r_burn}"] = f"=IF({col}{r_ebitda}<0,-{col}{r_ebitda},0)"
    ws3[f"{col}{r_cac}"] = f"={col}{r_pay}*0.15*{UE['CAC']}"
    ws3[f"{col}{r_ltv}"] = f"={UE['LTV']}"
    ws3[f"{col}{r_ltvcac}"] = f"={UE['LTV_CAC']}"
    ws3[f"{col}{r_payback}"] = f"={UE['Payback']}"
    ws3[f"{col}{r_be}"] = f"={col}{r_opex}/{UE['GM_per_user']}"

    for rr in [r_mau, r_pay, r_mrr, r_arr, r_cogs, r_gp, r_opex, r_ebitda, r_burn, r_cac, r_ltv, r_be]:
        ws3.cell(row=rr, column=2 + i).number_format = "#,##0"
    for rr in [r_gm, r_ebitdam]:
        ws3.cell(row=rr, column=2 + i).number_format = "0.0%"
    ws3.cell(row=r_ltvcac, column=2 + i).number_format = "0.00"
    ws3.cell(row=r_payback, column=2 + i).number_format = "0.0"

for rr in range(start_row, start_row + len(metric_rows)):
    for cc in range(1, len(headers) + 1):
        ws3.cell(row=rr, column=cc).border = BORDER
        if cc > 1:
            ws3.cell(row=rr, column=cc).font = CALC_FONT

autosize(ws3, [55] + [14] * len(mau_values))
ws3.freeze_panes = "B4"

note_row = start_row + len(metric_rows) + 1
ws3.cell(row=note_row, column=1,
         value="Break-even MAU здесь — это платящие пользователи, нужные чтобы Gross Profit = Opex ПРИ Opex ЭТОГО столбца (Opex растёт со шкалой инфры/команды, поэтому это не единая точка, а per-scenario ориентир; см. лист 'Break-Even' для динамической кривой).")
ws3.cell(row=note_row, column=1).font = NOTE_FONT
ws3.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=len(headers))

# chart: MRR & EBITDA vs MAU
chart = LineChart()
chart.title = "MRR vs EBITDA по сценариям MAU"
chart.y_axis.title = "USD/мес"
chart.x_axis.title = "MAU"
data = Reference(ws3, min_col=2, max_col=1 + len(mau_values), min_row=start_row + 2, max_row=start_row + 2)
data2 = Reference(ws3, min_col=2, max_col=1 + len(mau_values), min_row=start_row + 8, max_row=start_row + 8)
cats = Reference(ws3, min_col=2, max_col=1 + len(mau_values), min_row=hr, max_row=hr)
chart.add_data(data, titles_from_data=False)
chart.add_data(data2, titles_from_data=False)
chart.series[0].tx = openpyxl.chart.series.SeriesLabel(v="MRR")
chart.series[1].tx = openpyxl.chart.series.SeriesLabel(v="EBITDA")
chart.set_categories(cats)
chart.height = 9
chart.width = 20
ws3.add_chart(chart, f"B{note_row + 3}")

# =====================================================================
# SHEET 4: Break-Even (dynamic, static Opex baseline vs paying users)
# =====================================================================
ws4 = wb.create_sheet("Break-Even")
ws4["A1"] = "Точка безубыточности — статическая модель (Opex зафиксирован на уровне MVP-команды S2)"
ws4["A1"].font = TITLE_FONT
ws4.merge_cells("A1:D1")
ws4["A3"] = "Fixed Opex (MVP baseline, USD/mo)"
ws4["B3"] = (
    f"={A['Team cost (S1: 1 AI-first разработчик + fractional legal/design)']}"
    f"+{A['Infra fixed (hosting, DB, CDN, monitoring, GPU reserved)']}"
    f"+{A['Content moderation / compliance tooling']}"
    f"+{A['Office/admin/legal overhead']}"
)
ws4["A4"] = "Gross margin per paying user (USD/mo)"
ws4["B4"] = f"={UE['GM_per_user']}"
ws4["A5"] = "Break-even paying users"
ws4["B5"] = "=B3/B4"
ws4["A6"] = "Break-even MAU (при текущем conversion rate)"
ws4["B6"] = f"=B5/{A['% free users who convert to any paid tier /mo']}"
for rr in [3, 4, 5, 6]:
    ws4.cell(row=rr, column=2).font = CALC_FONT
    ws4.cell(row=rr, column=1).border = BORDER
    ws4.cell(row=rr, column=2).border = BORDER
ws4["B3"].number_format = "#,##0"
ws4["B4"].number_format = "0.00"
ws4["B5"].number_format = "#,##0"
ws4["B6"].number_format = "#,##0"
autosize(ws4, [45, 16])

# =====================================================================
# SHEET 5: Cash Flow (24-month monthly projection, S-curve growth)
# =====================================================================
ws5 = wb.create_sheet("Cash Flow 24mo")
ws5["A1"] = "Помесячный Cash Flow — 24 месяца (рост MAU задаётся кривой ниже, редактируемо)"
ws5["A1"].font = TITLE_FONT
ws5.merge_cells("A1:D1")

ws5["A3"] = "Starting cash (seed)"
ws5["B3"] = f"={A['Стартовый капитал (seed)']}"
ws5["B3"].font = CALC_FONT

hdr_row = 5
ws5.cell(row=hdr_row, column=1, value="Месяц")
style_header(ws5, hdr_row, 1, 25)
months = list(range(1, 25))
# S-curve target MAU milestones (editable input row) — user can overwrite any month's MAU directly.
target_curve = [100, 250, 500, 800, 1200, 1800, 2600, 3600, 5000, 6500, 8200, 10000,
                 13000, 16500, 20500, 25000, 30500, 36500, 43000, 50000, 58000, 67000, 78000, 100000]

for i, m in enumerate(months):
    ws5.cell(row=hdr_row, column=2 + i, value=m)

rows5 = ["MAU (input — редактируемо)", "Paying users", "MRR", "COGS", "Gross Profit",
         "Opex", "EBITDA", "Cumulative Cash"]
r0 = hdr_row + 1
for i, label in enumerate(rows5):
    ws5.cell(row=r0 + i, column=1, value=label).border = BORDER

for i, m in enumerate(months):
    col = get_column_letter(2 + i)
    ws5[f"{col}{r0}"] = target_curve[i]
    ws5[f"{col}{r0}"].fill = INPUT_FILL
    ws5[f"{col}{r0}"].font = INPUT_FONT
    ws5[f"{col}{r0+1}"] = f"={col}{r0}*{A['% free users who convert to any paid tier /mo']}"
    ws5[f"{col}{r0+2}"] = f"={col}{r0+1}*{UE['ARPU']}"
    ws5[f"{col}{r0+3}"] = f"={col}{r0+1}*{UE['COGS_per_user']}"
    ws5[f"{col}{r0+4}"] = f"={col}{r0+2}-{col}{r0+3}"
    ws5[f"{col}{r0+5}"] = (
        f"={A['Team cost (S1: 1 AI-first разработчик + fractional legal/design)']}"
        f"+{A['Infra fixed (hosting, DB, CDN, monitoring, GPU reserved)']}"
        f"+{A['Infra fixed scaling step (+ per 10k users)']}*({col}{r0}/10000)"
        f"+{A['Content moderation / compliance tooling']}"
        f"+{A['Office/admin/legal overhead']}"
    )
    ws5[f"{col}{r0+6}"] = f"={col}{r0+4}-{col}{r0+5}"
    if i == 0:
        ws5[f"{col}{r0+7}"] = f"=B3+{col}{r0+6}"
    else:
        prev = get_column_letter(2 + i - 1)
        ws5[f"{col}{r0+7}"] = f"={prev}{r0+7}+{col}{r0+6}"
    for rr in range(r0 + 1, r0 + 8):
        ws5.cell(row=rr, column=2 + i).number_format = "#,##0"
        if rr != r0:
            ws5.cell(row=rr, column=2 + i).font = CALC_FONT
        ws5.cell(row=rr, column=2 + i).border = BORDER
    ws5.cell(row=r0, column=2 + i).border = BORDER

autosize(ws5, [30] + [10] * 24)
ws5.freeze_panes = "B6"

chart2 = LineChart()
chart2.title = "Cumulative Cash (24 мес.)"
chart2.y_axis.title = "USD"
data3 = Reference(ws5, min_col=2, max_col=25, min_row=r0 + 7, max_row=r0 + 7)
cats2 = Reference(ws5, min_col=2, max_col=25, min_row=hdr_row, max_row=hdr_row)
chart2.add_data(data3, titles_from_data=False)
chart2.series[0].tx = openpyxl.chart.series.SeriesLabel(v="Cumulative Cash")
chart2.set_categories(cats2)
chart2.height = 9
chart2.width = 24
ws5.add_chart(chart2, f"B{r0 + 10}")

ws5.cell(row=r0 + 9, column=1,
         value="MAU-кривая — иллюстративный сценарий S-curve до 100k MAU за 24 мес.; каждая жёлтая ячейка редактируема для стресс-теста.")
ws5.cell(row=r0 + 9, column=1).font = NOTE_FONT

# =====================================================================
wb.save("financial-model.xlsx")
print("Saved financial-model.xlsx")
